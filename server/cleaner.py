import pandas as pd
import numpy as np
import pdfplumber
import io
import re
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import traceback


# ==========================================
# 0. CONFIGURATION & GLOBAL HELPERS
# ==========================================

# 1. The Headers for Excel Output (Friendly Names)
MANDATORY_HEADERS = [
    'Shift Date', 'Trip ID', 'Employee ID', 'Gender', 'EMP_CATEGORY', 'Employee Name', 
    'Shift Time', 'Pickup Time', 'Drop Time', 'Trip Direction', 'Cab Reg No', 
    'Cab Type', 'Vendor', 'Office', 'Airport Name', 'Landmark', 'Address', 
    'Flight Number', 'Flight Category', 'Flight Route', 'Flight Type', 
    'Trip Date', 'MiS Remark', 'In App/ Extra', 'UNA2', 'UNA', 
    'Route Status', 'Clubbing Status', 'GPS TIME', 'GPS REMARK',
    'passenger_mobile', 'driver_name', 'DRIVER_MOBILE' 
]

# 2. Map Friendly Headers -> SQLModel Fields (snake_case)
MANDATORY_DB_MAP = {
    'Shift Date': 'shift_date', 'Trip ID': 'trip_id', 'Employee ID': 'employee_id', 
    'Gender': 'gender', 'EMP_CATEGORY': 'emp_category', 'Employee Name': 'employee_name', 
    'Shift Time': 'shift_time', 'Pickup Time': 'pickup_time', 'Drop Time': 'drop_time', 
    'Trip Direction': 'trip_direction', 'Cab Reg No': 'cab_reg_no', 'Cab Type': 'cab_type', 
    'Vendor': 'vendor', 'Office': 'office', 'Airport Name': 'airport_name', 
    'Landmark': 'landmark', 'Address': 'address', 'Flight Number': 'flight_number', 
    'Flight Category': 'flight_category', 'Flight Route': 'flight_route', 
    'Flight Type': 'flight_type', 'Trip Date': 'trip_date', 'MiS Remark': 'mis_remark', 
    'In App/ Extra': 'in_app_extra', 'UNA2': 'una2', 'UNA': 'una', 
    'Route Status': 'route_status', 'Clubbing Status': 'clubbing_status', 
    'GPS TIME': 'gps_time', 'GPS REMARK': 'gps_remark',
    'passenger_mobile': 'passenger_mobile', 'driver_name': 'driver_name', 'DRIVER_MOBILE': 'driver_mobile'
}

def standardize_dataframe(df):
    """
    Standardizes DataFrame to match SQLModel definitions:
    1. Generates 'unique_id' (Composite Key).
    2. Ensures ALL mapped columns exist (fills missing with "").
    3. Returns DataFrame with [Mandatory Columns] + [Any Extra Columns].
    """
    if df is None or df.empty:
        return None

    # 1. Generate Unique ID
    if 'trip_id' in df.columns and 'employee_id' in df.columns:
        df['unique_id'] = df['trip_id'].astype(str).str.strip() + df['employee_id'].astype(str).str.strip()
        # Filter invalid IDs
        df = df[
            df["unique_id"].notna() & 
            (df["unique_id"] != "") & 
            (~df["unique_id"].str.contains("nan|None", case=False))
        ]
    else:
        return None # Critical missing data

    # 2. Sync Columns with SQLModel (Fill Missing)
    target_cols = list(MANDATORY_DB_MAP.values())
    for col in target_cols:
        if col not in df.columns:
            df[col] = ""

    # 3. Populate specific logic columns
    df['una2'] = df['unique_id'] # As requested
    # Optional: populate 'una' if needed, or leave blank

    # 4. Organize Columns: Mandatory -> Extra -> unique_id
    current_cols = df.columns.tolist()
    reserved_cols = set(target_cols) | {'unique_id'}
    extra_cols = [c for c in current_cols if c not in reserved_cols]
    
    final_order = target_cols + extra_cols + ['unique_id']
    return df[final_order]

def get_xls_style_data(book, xf_index):
    """ Helper to extract style (Hex Color, Bold) from xlrd book. """
    try:
        xf = book.xf_list[xf_index]
        font = book.font_list[xf.font_index]
        
        def get_hex(idx):
            if idx is not None and idx < 64:
                rgb = book.colour_map.get(idx)
                if rgb and rgb != (0,0,0) and rgb != (255, 255, 255):
                    return "{:02x}{:02x}{:02x}".format(*rgb).upper()
            return None

        bg_hex = get_hex(xf.background.pattern_colour_index)
        font_hex = get_hex(font.colour_index)
        return bg_hex, font_hex, bool(font.bold)
    except:
        return None, None, False

def create_styled_excel(df, filename_prefix="Cleaned"):
    """ Generates Excel with consistent formatting. """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Invert map to get Friendly Headers
        inv_map = {v: k for k, v in MANDATORY_DB_MAP.items()}
        df_export = df.rename(columns=inv_map)
        
        # Order: Mandatory Headers -> Extra Headers
        export_cols = [h for h in MANDATORY_HEADERS if h in df_export.columns]
        remaining = [c for c in df_export.columns if c not in MANDATORY_HEADERS and c != 'unique_id']
        export_cols += remaining
        
        df_export = df_export[export_cols]
        
        df_export.to_excel(writer, index=False, sheet_name='Raw_Data')
        
        # Apply Styles
        workbook = writer.book
        worksheet = writer.sheets['Raw_Data']
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#0070C0', 'font_color': 'white'})
        
        for i, col in enumerate(df_export.columns):
            worksheet.write(0, i, col, header_fmt)
            worksheet.set_column(i, i, 15)
            
    output.seek(0)
    return df, output, f"{filename_prefix}.xlsx"


# ==========================================
# 1. CLIENT DATA CLEANER
# ==========================================
def process_client_data(file_content):
    try:
        # Columns to Remove based on your request
        DROP_COLS = [
            'Bunit ID', 'Cycle Start', 'Cycle End', 'Billing Period', 'Project', 'Cost Center', 
            'Department', 'Planned Emp Count', 'Travelled Emp Count', 'Billable Emp Count', 
            'No Show', 'Planned Escort', 'Actual Escort', 'Emp Km', 'Trip Cost', 
            'Trip AC Cost', 'Per Emp Cost', 'Escort Cost', 'Penalty', 'Vendor Penalty', 
            'Total Cost', 'Assigned Contract', 'Cab Contract', 'Billing Zone', 
            'Trip Billing Zone', 'Emp Sigin Type', 'Escort ID', 'Toll Cost', 
            'State Tax Cost', 'Parking Or Toll Cost', 'Per Employee Overhead Cost', 
            'Trip Source', 'Extra Kms Based On Billable Employee Count', 'Billing Kms', 
            'Actual Kms At Employee Level', 'Grid Km', 'Employee Adjustment Distance', 
            'Trip Adjustment', 'Total Distance'
        ]

        # Input Mapping
        COL_MAP = {
            "Shift Date": "shift_date", "Trip ID": "trip_id", "Employee ID": "employee_id", 
            "Gender": "gender", "Employee Name": "employee_name", "Shift Time": "shift_time", 
            "Pickup Time": "pickup_time", "Drop Time": "drop_time", "Trip Direction": "trip_direction", 
            "Cab Reg No": "cab_reg_no", "Cab Type": "cab_type", "Vendor": "vendor", 
            "Office": "office", "Airport Name": "airport_name", "Landmark": "landmark", 
            "Address": "address", "Flight Number": "flight_number", "Flight Category": "flight_category", 
            "Flight Route": "flight_route", "Flight Type": "flight_type"
        }
        
        df = pd.read_excel(io.BytesIO(file_content), dtype=str)
        
        # 1. Drop Unwanted
        df = df.drop(columns=[c for c in DROP_COLS if c in df.columns], errors='ignore')

        # 2. Data Cleaning
        if "Cab Reg No" in df.columns:
            df["Cab Reg No"] = df["Cab Reg No"].str.replace("-", "", regex=False).str.upper()

        if "Trip Direction" in df.columns:
            df["Trip Direction"] = df["Trip Direction"].astype(str).str.strip().str.title().replace({
                "Login": "Pickup", "Logout": "Drop"
            })

        # 3. Rename to DB Columns
        df_db = df.rename(columns=COL_MAP).fillna("")
        
        # 4. Standardize (Matches ClientData Model)
        df_db = standardize_dataframe(df_db)
        if df_db is None: return None, None, None

        return create_styled_excel(df_db, "Client_Cleaned")
    except Exception as e:
        print(f"❌ Client Cleaner Error: {e}")
        return None, None, None


 # Added for debugging

# ==========================================
# 2. RAW DATA CLEANER
# ==========================================
def _clean_single_raw_df(df):
    try:

        df = df.replace({np.nan: None, "nan": None})
        # Trip ID Logic
        df["Trip_ID"] = np.where(df.iloc[:, 10].astype(str).str.startswith("T"), df.iloc[:, 10], np.nan)
        df["Trip_ID"] = df["Trip_ID"].ffill()

        # Identify Row Types
        is_header = df.iloc[:, 1].astype(str).str.contains("UNITED FACILITIES", na=False)
        # DEBUG: Relaxed regex slightly to ensure we catch rows even if there are formatting oddities
        is_passenger = df.iloc[:, 0].astype(str).str.match(r"^[0-9]+$") 

        # Extraction Maps
        h_map = {0: 'TRIP_DATE', 1: 'AGENCY_NAME', 2: 'D_LOGIN', 3: 'VEHICLE_NO', 4: 'DRIVER_NAME', 6: 'DRIVER_MOBILE', 7: 'MARSHALL', 8: 'DISTANCE', 9: 'EMP_COUNT', 10: 'TRIP_COUNT'}
        p_map = {0: 'PAX_NO', 1: 'REPORTING_TIME', 2: 'EMPLOYEE_ID', 3: 'EMPLOYEE_NAME', 4: 'GENDER', 5: 'EMP_CATEGORY', 6: 'FLIGHT_NO.', 7: 'ADDRESS', 8: 'REPORTING_LOCATION', 9: 'LANDMARK', 10: 'PASSENGER_MOBILE'}

        df_h = df[is_header].rename(columns=h_map)
        df_p = df[is_passenger].rename(columns=p_map)
        
        # DEBUG CHECK
        if df_h.empty or df_p.empty:
            print("DEBUG: Header or Passenger dataframe is empty. Check regex or file format.")

        cols_h = [c for c in h_map.values() if c in df_h.columns] + ['Trip_ID']
        cols_p = [c for c in p_map.values() if c in df_p.columns] + ['Trip_ID']
        
        merged = pd.merge(df_p[cols_p], df_h[cols_h], on='Trip_ID', how='left')

        # Cleaning Logic
        for col in ['TRIP_ID', 'Trip_ID']:
            if col in merged.columns:
                merged[col] = merged[col].astype(str).str.replace('T', '', regex=False)
                merged.rename(columns={col: 'TRIP_ID'}, inplace=True)

        if 'AGENCY_NAME' in merged.columns:
            merged['AGENCY_NAME'] = merged['AGENCY_NAME'].apply(lambda x: "UNITED FACILITIES" if "UNITED FACILITIES" in str(x).upper() else x)

        if 'VEHICLE_NO' in merged.columns:
            merged['VEHICLE_NO'] = merged['VEHICLE_NO'].astype(str).str.replace('-', '', regex=False)

        if 'D_LOGIN' in merged.columns:
            login = merged['D_LOGIN'].astype(str).str.strip().str.split(' ', n=1, expand=True)
            if len(login.columns) > 0: merged['DIRECTION'] = login[0].str.upper().replace({'LOGIN': 'PICKUP', 'LOGOUT': 'DROP'})
            if len(login.columns) > 1: merged['SHIFT_TIME'] = login[1]

        # Explicit Removal
        cols_to_remove = ['PAX_NO', 'D_LOGIN', 'MARSHALL', 'DISTANCE', 'EMP_COUNT', 'TRIP_COUNT']
        merged = merged.drop(columns=cols_to_remove, errors='ignore')

        for col in merged.select_dtypes(include=['object']):
            merged[col] = merged[col].astype(str).str.upper().str.strip()

        return merged
    except Exception as e:
        # DEBUG: Print actual error
        print(f"Error in _clean_single_raw_df: {e}")
        traceback.print_exc()
        return pd.DataFrame()

def process_raw_data(file_list_bytes):
    all_dfs = []
    for filename, content in file_list_bytes: # Added filename to loop for better debug
        try:
            print(f"Processing file: {filename}") # DEBUG
            df_raw = pd.read_excel(io.BytesIO(content), header=None,dtype=str).dropna(how="all").reset_index(drop=True)
            cleaned = _clean_single_raw_df(df_raw)
            if not cleaned.empty: 
                all_dfs.append(cleaned)
            else:
                print(f"Warning: File {filename} resulted in empty data.")
        except Exception as e:
            # DEBUG: Print actual error
            print(f"FAILED processing file {filename}: {e}")
            traceback.print_exc()
            pass

    if not all_dfs: 
        print("No valid dataframes found in any files.")
        return None, None, None
        
    final_df = pd.concat(all_dfs, ignore_index=True)

    # Input Map -> DB Columns
    DB_MAP = {
        'TRIP_DATE': 'shift_date', 'TRIP_ID': 'trip_id', 'AGENCY_NAME': 'vendor', 
        'FLIGHT_NO.': 'flight_number', 'EMPLOYEE_ID': 'employee_id', 'EMPLOYEE_NAME': 'employee_name', 
        'GENDER': 'gender', 'EMP_CATEGORY': 'emp_category', 'ADDRESS': 'address', 
        'PASSENGER_MOBILE': 'passenger_mobile', 'LANDMARK': 'landmark', 'VEHICLE_NO': 'cab_reg_no',
        'DRIVER_NAME': 'driver_name', 'DRIVER_MOBILE': 'driver_mobile', 'DIRECTION': 'trip_direction',
        'SHIFT_TIME': 'shift_time', 'REPORTING_TIME': 'pickup_time', 'REPORTING_LOCATION': 'office'    
    }

    final_db = final_df.rename(columns=DB_MAP).fillna("")
    
    # 2. Date/Time Formatting
    if 'shift_date' in final_db.columns:
        final_db['shift_date'] = pd.to_datetime(final_db['shift_date'], errors='coerce').dt.strftime('%d-%m-%Y')
        final_db['trip_date'] = final_db['shift_date'] # Duplicate to trip_date
    
    if 'shift_time' in final_db.columns:
        final_db['shift_time'] = pd.to_datetime(final_db['shift_time'], errors='coerce', format='mixed').dt.strftime('%H:%M')

    # 3. Standardize (Matches RawTripData Model)
    # DEBUG: Wrapped in try/except in case this function is missing or failing
    try:
        final_db = standardize_dataframe(final_db)
    except NameError:
        print("CRITICAL: 'standardize_dataframe' function is not defined.")
    except Exception as e:
        print(f"Error in standardize_dataframe: {e}")
        traceback.print_exc()
        return None, None, None

    if final_db is None: return None, None, None

    # DEBUG: Wrapped in try/except
    try:
        return create_styled_excel(final_db, "Raw_Cleaned")
    except NameError:
        print("CRITICAL: 'create_styled_excel' function is not defined.")
        return final_db, None, None
# ==========================================
# 3. OPERATION DATA CLEANER
# ==========================================
def process_operation_data(file_list_bytes):
    # Mapping specifically for reading operation files
    READ_MAP = {
        'DATE': 'Shift Date', 'TRIP ID': 'Trip ID', 'FLT NO.': 'Flight Number', 'SAP ID': 'Employee ID',
        'EMP NAME': 'Employee Name', 'EMPLOYEE ADDRESS': 'Address', 'PICKUP LOCATION': 'Landmark',
        'DROP LOCATION': 'drop_location', 'CAB NO': 'Cab Reg No', 'AIRPORT DROP': 'Shift Time', 
        'GUARD ROUTE': 'guard_route', 'MARSHALL': 'guard_route', 'REMARKS': 'MiS Remark'
    }
    SKIP_HEADERS = ['PICKUP TIME', 'CONTACT NO', 'CONTACT NO.']

    wb = Workbook()
    ws = wb.active
    ws.title = "Operation_Data"
    
    # Styles
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write Mandatory Headers
    for col_idx, header in enumerate(MANDATORY_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill; cell.font = header_font; cell.alignment = align_center; cell.border = border
        ws.column_dimensions[cell.column_letter].width = 15

    target_row = 2
    files_processed = 0
    data_rows = [] 
    
    # Track extra headers found dynamically
    extra_headers_map = {} 
    next_extra_col_idx = len(MANDATORY_HEADERS) + 1

    for filename, content in file_list_bytes:
        if not filename.lower().endswith('.xls'): continue
        try:
            rb = xlrd.open_workbook(file_contents=content, formatting_info=True)
            rs = rb.sheet_by_index(0)
            
            # Map Source Columns
            source_headers = [str(rs.cell_value(0, c)).strip().upper() for c in range(rs.ncols)]
            col_to_target_map = {} 
            
            for idx, raw_header in enumerate(source_headers):
                if any(skip in raw_header for skip in SKIP_HEADERS): continue
                
                match = next((val for key, val in READ_MAP.items() if key in raw_header), None)
                if match:
                    col_to_target_map[idx] = {'type': 'mandatory', 'name': match}
                else:
                    clean_name = raw_header
                    if clean_name not in extra_headers_map:
                        extra_headers_map[clean_name] = next_extra_col_idx
                        # Add header to Excel
                        cell = ws.cell(row=1, column=next_extra_col_idx, value=clean_name)
                        cell.fill = header_fill; cell.font = header_font; cell.alignment = align_center; cell.border = border
                        ws.column_dimensions[cell.column_letter].width = 15
                        next_extra_col_idx += 1
                    col_to_target_map[idx] = {'type': 'extra', 'name': clean_name}

            start_row = 0 if files_processed == 0 else 1 
            
            for r_idx in range(1, rs.nrows):
                row_vals = rs.row_values(r_idx)
                # Skip empty/near empty rows
                if sum(1 for v in row_vals if str(v).strip() != "") <= 2: continue
                
                row_data_map = {} 
                db_row_dict = {}  
                
                for c_idx in range(rs.ncols):
                    if c_idx not in col_to_target_map: continue
                    mapping = col_to_target_map[c_idx]
                    target_header = mapping['name']
                    
                    val = rs.cell_value(r_idx, c_idx)
                    bg, fg, bold = get_xls_style_data(rb, rs.cell_xf_index(r_idx, c_idx))
                    
                    style_data = {'val': val, 'bg': bg, 'fg': fg, 'bold': bold}
                    
                    if mapping['type'] == 'mandatory':
                         row_data_map[target_header] = style_data
                         if target_header in MANDATORY_DB_MAP:
                             db_row_dict[MANDATORY_DB_MAP[target_header]] = val
                         elif target_header == 'guard_route':
                             db_row_dict['guard_route'] = val
                         elif target_header == 'drop_location':
                             db_row_dict['drop_location'] = val
                    else:
                        row_data_map[f"EXTRA_{target_header}"] = style_data
                        db_row_dict[target_header] = val # Store extra in DB dict too if needed by model

                # Write Row to Excel
                ws.row_dimensions[target_row].height = 25
                
                # Write Mandatory
                for c_out, m_header in enumerate(MANDATORY_HEADERS, 1):
                    cell = ws.cell(row=target_row, column=c_out)
                    if m_header in row_data_map:
                        data = row_data_map[m_header]
                        cell.value = data['val']
                        if data['bg']: cell.fill = PatternFill(start_color=data['bg'], end_color=data['bg'], fill_type='solid')
                        cell.font = Font(size=11, bold=data['bold'], color=data['fg'] if data['fg'] else None)
                    else: cell.value = ""
                    cell.alignment = align_center; cell.border = border
                
                # Write Extra
                for extra_name, extra_col_idx in extra_headers_map.items():
                    key = f"EXTRA_{extra_name}"
                    cell = ws.cell(row=target_row, column=extra_col_idx)
                    if key in row_data_map:
                        data = row_data_map[key]
                        cell.value = data['val']
                        if data['bg']: cell.fill = PatternFill(start_color=data['bg'], end_color=data['bg'], fill_type='solid')
                        cell.font = Font(size=11, bold=data['bold'], color=data['fg'] if data['fg'] else None)
                    else: cell.value = ""
                    cell.alignment = align_center; cell.border = border
                
                data_rows.append(db_row_dict)
                target_row += 1
            
            files_processed += 1
            rb.release_resources()
        except Exception as e: print(f"Error {filename}: {e}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    if not data_rows: return None, output, "Operation_Cleaned.xlsx"

    df_db = pd.DataFrame(data_rows)

    # Post-Process Dates/Times
    if 'shift_date' in df_db.columns: 
        df_db['shift_date'] = pd.to_datetime(df_db['shift_date'], errors='coerce').dt.strftime('%d-%m-%Y')
    if 'shift_time' in df_db.columns: 
        df_db['shift_time'] = pd.to_datetime(df_db['shift_time'], errors='coerce', format='mixed').dt.strftime('%H:%M')
    if 'guard_route' in df_db.columns:
        df_db['guard_route'] = df_db['guard_route'].astype(str).apply(lambda x: 'Guard' if 'marshall' in x.lower() else '')
        # Map internal 'guard_route' to 'route_status' or 'marshall' field if strictly required
        # Note: OperationData Model uses 'route_status', but raw 'marshall' map logic existed. 
        # Standardize will handle missing cols, but we populated 'guard_route' in dict.

    # Standardize (Matches OperationData Model)
    df_db = standardize_dataframe(df_db)
    return df_db, output, "Operation_Cleaned.xlsx"




# ==========================================
# 4. BAROW DATA CLEANER
# ==========================================
def process_ba_row_data(file_content):
    try:
        print("🔹 Starting BA Row Data Processing...")
        
        # 1. READ CSV
        df = pd.read_csv(io.BytesIO(file_content), low_memory=False)
        print(f"🔹 CSV Loaded. Columns: {list(df.columns[:5])}...")

        df.columns = df.columns.str.strip()
        print(f"🔹 CSV Loaded. Found Columns: {list(df.columns)}")

        # 2. FILTER & TRANSFORM
        if "Trip Id" in df.columns:
            df["Trip Id"] = pd.to_numeric(df["Trip Id"], errors='coerce').fillna(0)
            df = df[df["Trip Id"] != 0]

        # Shift Time Logic
        if "Trip Type" in df.columns:
            df["Shift Time"] = df["Trip Type"].astype(str)
            mask_log = df["Shift Time"].str.contains("LOGIN|LOGOUT", case=False, na=False)
            df.loc[mask_log, "Shift Time"] = "00:00"
        else:
            df["Shift Time"] = "00:00"

        # Trip Direction
        if "Direction" in df.columns:
            df["Trip Direction"] = df["Direction"].str.upper().map({
                "LOGIN": "PICKUP", "LOGOUT": "DROP"
            })
        else:
            df["Trip Direction"] = ""

        # Safe Column Access for Pickup/Drop
        df["Pickup Time"] = df.get("Duty Start", "")
        df["Drop Time"] = df.get("Duty End", "")

        # Registration Cleaning
        if "Registration" in df.columns:
            df["Registration"] = (
                df["Registration"].astype(str)
                .str.replace("-", "", regex=False)
                .str.replace(" ", "", regex=False)
                .str.upper()
            )

        # Location Logic
        is_drop = df["Trip Direction"] == "DROP"
        is_pickup = df["Trip Direction"] == "PICKUP"
        
        start_addr = df.get("Start Location Address", "")
        end_addr = df.get("End Location Address", "")
        start_land = df.get("Start Location Landmark", "")
        end_land = df.get("End Location Landmark", "")

        df["Airport Name"] = np.where(is_drop, start_addr, end_addr)
        df["Address"] = np.where(is_pickup, start_addr, end_addr)
        df["Landmark"] = np.where(is_pickup, start_land, end_land)

        if "Leg Date" in df.columns:
            # Standard case
            df["Trip Date"] = df["Leg Date"].astype(str) + " " + df["Shift Time"].astype(str)
        elif "Date" in df.columns:
            # Fallback to 'Date' column
            df["Trip Date"] = df["Date"].astype(str) + " " + df["Shift Time"].astype(str)
        elif "Pickup Time" in df.columns:
            # Fallback: Extract Date from Duty Start (e.g., '2026-01-01 18:00:00')
            print("⚠️ 'Leg Date' missing. Extracting date from 'Pickup Time'.")
            df["Temp_Date"] = pd.to_datetime(df["Pickup Time"], errors='coerce').dt.strftime('%d-%m-%Y')
            df["Trip Date"] = df["Temp_Date"].astype(str) + " " + df["Shift Time"].astype(str)
            df["Leg Date"] = df["Temp_Date"] # Fill Leg Date so it's not empty in DB
        else:
            # Worst case: No date found
            print("❌ ERROR: Could not find 'Leg Date', 'Date', or 'Pickup Time'. Trip Date will be empty.")
            df["Leg Date"] = ""
            df["Trip Date"] = ""

        if "Trip Date" in df.columns:
            df["Shift Date"] = df["Trip Date"]
        else:
            df["Shift Date"] = ""

        df["In App/ Extra"] = "BA Row Data"
        df["BA REMARK"] = df.get("Trip Status", "")
        df["MiS Remark"] = df.get("Comments", "")

        # 3. PREPARE DATABASE MAPPING (Title Case -> snake_case)
        DB_MAP = {
            "Leg Date": "leg_date",
            "Trip Id": "trip_id",
            "Employee ID": "employee_id",
            "Gender": "gender",
            "EMP_CATEGORY": "emp_category",
            "Employee Name": "employee_name",
            "Shift Time": "shift_time",
            "Pickup Time": "pickup_time",
            "Drop Time": "drop_time",
            "Trip Direction": "trip_direction",
            "Registration": "cab_reg_no",
            "Cab Type": "cab_type",
            "Vendor": "vendor",
            "Office": "office",
            "Airport Name": "airport_name",
            "Landmark": "landmark",
            "Address": "address",
            "Flight Number": "flight_number",
            "Flight Category": "flight_category",
            "Flight Route": "flight_route",
            "Flight Type": "flight_type",
            "Trip Date": "trip_date",
            "MiS Remark": "mis_remark",
            "In App/ Extra": "in_app_extra",
            "Traveled Employee Count": "traveled_emp_count",
            "UNA2": "una2",
            "UNA": "una",
            "BA REMARK": "ba_remark",
            "Route Status": "route_status",
            "Clubbing Status": "clubbing_status",
            "GPS TIME": "gps_time",
            "GPS REMARK": "gps_remark",
            "Billing Zone Name": "billing_zone_name",
            "Leg Type": "leg_type",
            "Trip Source": "trip_source",
            "Trip Type": "trip_type",
            "Leg Start": "leg_start",
            "Leg End": "leg_end",
            "Audit Results": "audit_results",
            "Audit Done By": "audit_done_by",
            "Trip Audited": "trip_audited"
        }

        # 4. FIX: ENSURE ALL COLUMNS EXIST
        # This loop prevents the KeyError by creating missing columns
        for col in DB_MAP.keys():
            if col not in df.columns:
                df[col] = ""

        # 5. SELECT AND RENAME
        # Now it is safe to select because we guaranteed they exist
        df_final = df[list(DB_MAP.keys())].copy()
        df_final.rename(columns=DB_MAP, inplace=True)

        print(f"🔹 Data Transformed. Renamed columns to: {list(df_final.columns[:5])}...")
        print("🔹 Calling Standardizer...")

        # 6. STANDARDIZE
        if 'standardize_dataframe' in globals():
            df_final = standardize_dataframe(df_final)
            
            if df_final is None:
                print("❌ Standardization failed. Check if DB model matches these columns.")
                return None, None, None
        else:
            print("⚠️ 'standardize_dataframe' function not found. Skipping.")

        # 7. EXPORT
        print("🔹 Generating Excel...")
        return create_styled_excel(df_final, "BA_Row_Data_Cleaned")

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ BA Row Data Cleaner Error: {e}")
        return None, None, None


# ==========================================
# 4. FASTAG DATA CLEANER (PDF) - MULTI FILE
# ==========================================

# ==========================================
# HELPER: ICICI SPECIFIC CLEANER
# ==========================================
def _process_icici(pdf_obj):
    all_rows = []
    for page in pdf_obj.pages:
        tables = page.extract_tables()
        for table in tables:
            if table:
                all_rows.extend(table)
    
    if not all_rows: return pd.DataFrame()

    df = pd.DataFrame(all_rows)

    # 1. Find Header (Look for "Date & Time" or "Transaction Description")
    header_idx = -1
    for i in range(min(20, len(df))):
        row_str = " ".join([str(x).lower() for x in df.iloc[i] if x])
        if "date" in row_str and "description" in row_str:
            header_idx = i
            break
    
    if header_idx == -1: return pd.DataFrame()

    # 2. Set Header & Remove Empty Columns
    df.columns = df.iloc[header_idx]
    df = df.iloc[header_idx+1:].reset_index(drop=True)
    df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]
    df = df.loc[:, df.columns != 'nan'] # Remove empty columns
    
    # 3. Rename to User's Desired Format
    col_map = {}
    for col in df.columns:
        c_low = col.lower()
        if "unique" in c_low and "id" in c_low:
            col_map[col] = "Unique Transaction ID"
        elif "date" in c_low:
            col_map[col] = "Travel Date Time"
        elif "activity" in c_low:
            col_map[col] = "Activity"
        elif "description" in c_low: 
            col_map[col] = "Plaza Name"  # User requested this mapping for ICICI
        elif "vehicle" in c_low:
            col_map[col] = "Vehicle No"
        elif "amount" in c_low and "dr" in c_low:
            col_map[col] = "Tag Dr/Cr"
        elif "plaza" in c_low and "id" in c_low:
            col_map[col] = "Plaza ID"

    df.rename(columns=col_map, inplace=True)
    return df

import pdfplumber
import pandas as pd
import io
import re
import numpy as np

# ==========================================
# HELPER: CLEANING UTILS
# ==========================================
def _clean_columns(columns):
    """Standardizes column names to snake_case (User's Logic)"""
    cleaned = (
        columns
        .astype(str)
        .str.replace(r"\n", " ", regex=True)        # remove line breaks
        .str.replace(r"\t", " ", regex=True)        # remove tabs
        .str.replace(r"\s+", " ", regex=True)       # normalize spaces
        .str.strip()                                # trim edges
        .str.lower()                                # lowercase
        .str.replace(r"[^\w\s]", "", regex=True)    # remove special chars
        .str.replace(" ", "_")                      # snake_case
    )
    return cleaned

def _clean_cell_value(x):
    """Normalizes spaces and handles None/NaN"""
    if isinstance(x, str):
        x = x.replace("\n", " ").replace("\t", " ")
        x = re.sub(r"\s+", " ", x).strip()
        if x.lower() in ["na", "n/a", "null", "none", ""]:
            return np.nan
        return x
    return x

# ==========================================
# HELPER: ICICI SPECIFIC CLEANER (YOUR PERFECT CODE)
# ==========================================
def _process_icici(pdf_obj):
    all_tables = []
    
    # 1. Extract Tables
    for page in pdf_obj.pages:
        tables = page.extract_tables()
        for table in tables:
            if table:
                all_tables.append(pd.DataFrame(table))
    
    if not all_tables: return pd.DataFrame()

    # 2. Merge
    df = pd.concat(all_tables, ignore_index=True)

    # 3. Hardcoded Drop (0-11) as requested
    # We check length first to avoid errors on empty files
    if len(df) > 12:
        df = df.drop(index=[0,1,2,3,4,5,6,7,8,9,10,11]).reset_index(drop=True)
    else:
        return pd.DataFrame()

    # 4. Set Header
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True) # Remove header row from data

    # 5. Clean Columns
    df.columns = _clean_columns(df.columns)
    df.columns = df.columns.str.strip()
    
    # Specific rename from your script
    df = df.rename(columns={"date__time": "date_time"})

    # 6. Extract Vehicle Number
    # Logic: First row, date_time column, split by space
    if not df.empty and "date_time" in df.columns:
        try:
            raw_val = str(df.loc[0, "date_time"])
            vehicle_no = raw_val.split(" ")[0]
            df["vehicle_no"] = vehicle_no
            
            # Remove that row used for extraction
            df = df.drop(index=[0]).reset_index(drop=True)
        except:
            df["vehicle_no"] = ""
    else:
        df["vehicle_no"] = ""

    # 7. Add Plaza ID placeholder
    df["plaza_id"] = ""

    # 8. Standardize Column Names (Replacements)
    replacements = {
        "drcr": "debit_credit",
        "rscr": "rupees_credit",
        "rsdr": "rupees_debit",
        "rs": "rupees",
        "amt": "amount",
        "bal": "balance"
    }
    for k, v in replacements.items():
        df.columns = df.columns.str.replace(k, v, regex=False)

    # 9. Drop unwanted columns
    df = df.drop(columns=["nan", "amount_rupees_credit"], errors="ignore")

    # 10. Final Rename Map
    rename_map = {
        "transaction_description": "plaza_name",
        "date_time": "travel_date_time",
        "vehicle_no": "vehicle_number",
        "amount_rupees_debit": "tag_debit_credit"
    }
    df = df.rename(columns=rename_map)

    # 11. Drop Empty Rows based on critical columns
    subset_cols = [
        "vehicle_number", "travel_date_time", "unique_transaction_id",
        "plaza_name", "activity", "tag_debit_credit"
    ]
    # Only drop if columns exist
    existing_subset = [c for c in subset_cols if c in df.columns]
    if existing_subset:
        df = df.dropna(subset=existing_subset)

    # 12. Final Column Selection
    final_columns = [
        "vehicle_number",
        "travel_date_time",
        "unique_transaction_id",
        "plaza_name",
        "plaza_id",
        "activity",
        "tag_debit_credit"
    ]
    
    # Add missing columns if any
    for col in final_columns:
        if col not in df.columns:
            df[col] = ""

    df = df[final_columns]

    # 13. Filter out repeated headers
    if "plaza_name" in df.columns:
        df = df[df["plaza_name"].astype(str).str.contains("transaction description", case=False, na=False) == False]

    # Map to Title Case for Final Output consistency
    final_title_map = {
        "vehicle_number": "Vehicle No",
        "travel_date_time": "Travel Date Time",
        "unique_transaction_id": "Unique Transaction ID",
        "plaza_name": "Plaza Name",
        "plaza_id": "Plaza ID",
        "activity": "Activity",
        "tag_debit_credit": "Tag Dr/Cr"
    }
    df.rename(columns=final_title_map, inplace=True)

    return df

# ==========================================
# HELPER: IDFC SPECIFIC CLEANER (YOUR PREVIOUS PERFECT CODE)
# ==========================================
# ==========================================
# HELPER: CLEANING UTILS (From your logic)
# ==========================================
def _clean_columns(columns):
    """Standardizes column names to snake_case"""
    cleaned = (
        columns
        .astype(str)
        .str.replace(r"\n", " ", regex=True)
        .str.replace(r"\t", " ", regex=True)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        .str.lower()
        .str.replace(r"[^\w\s]", "", regex=True)
        .str.replace(" ", "_")
    )
    return cleaned

def _clean_cell_value(x):
    """Normalizes spaces and handles None/NaN"""
    if isinstance(x, str):
        x = x.replace("\n", " ").replace("\t", " ")
        x = re.sub(r"\s+", " ", x).strip()
        if x.lower() in ["na", "n/a", "null", "none", ""]:
            return np.nan
        return x
    return x

def _clean_datetime(x):
    """Fixes broken years (2 025) and time spacing"""
    if not isinstance(x, str):
        return x
    x = re.sub(r"\s+", " ", x).strip()
    # Fix broken year (2 025 -> 2025)
    x = re.sub(r"(\d{2})-(\d)\s(\d{3})", r"\1-\2\3", x)
    # Fix time spacing (23:3 2:46 -> 23:32:46)
    x = re.sub(r"(\d{2}):(\d)\s(\d):(\d{2})", r"\1:\2\3:\4", x)
    return x

def _clean_reference_id(x):
    if not isinstance(x, str):
        return x
    return x.replace(" ", "")

def _clean_vehicle_no(x):
    if isinstance(x, str):
        return x.replace(" ", "").strip()
    return x

# ==========================================
# HELPER: IDFC SPECIFIC CLEANER
# ==========================================
def _process_idfc(pdf_obj):
    all_tables = []
    for page in pdf_obj.pages:
        tables = page.extract_tables()
        for table in tables:
            if table:
                all_tables.append(pd.DataFrame(table))
    
    if not all_tables:
        print("⚠️ IDFC: No tables found.")
        return pd.DataFrame()

    df = pd.concat(all_tables, ignore_index=True)

    # 1. Drop known junk rows
    if len(df) > 5:
        df = df.drop(index=[0,1,2,3,4]).reset_index(drop=True)
    else:
        return pd.DataFrame()

    # 2. Set Headers
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)
    df.columns = _clean_columns(df.columns)

    # 3. Clean Headers & Values
    cols_to_drop = ["processed_date_time", "pool_drcr", "closing_pool_balance_rs", "closing_tag_balance_rs"]
    df = df.drop(columns=cols_to_drop, errors="ignore")

    replacements = {"drcr": "debit_credit", "rs": "rupees", "amt": "amount", "bal": "balance"}
    for k, v in replacements.items():
        df.columns = df.columns.str.replace(k, v, regex=False)

    for col in df.columns:
        df[col] = df[col].apply(_clean_cell_value)

    # 4. 🔥 REPAIR SPLIT ROWS (Merging wrapped IDs)
    if "travel_date_time" in df.columns and "unique_transaction_id" in df.columns:
        rows_to_drop = []
        for i in range(1, len(df)):
            curr_date = str(df.loc[i, "travel_date_time"])
            curr_id_frag = str(df.loc[i, "unique_transaction_id"])
            
            is_invalid_date = (curr_date == "" or curr_date.lower() == "nan" or "nan" in curr_date.lower())
            has_fragment = (curr_id_frag != "" and curr_id_frag.lower() != "nan")
            
            if is_invalid_date and has_fragment:
                prev_idx = i - 1
                while prev_idx in rows_to_drop and prev_idx > 0:
                    prev_idx -= 1
                
                if prev_idx >= 0:
                    current_val = str(df.loc[prev_idx, "unique_transaction_id"])
                    if "HR" not in curr_id_frag and "DL" not in curr_id_frag: 
                         df.at[prev_idx, "unique_transaction_id"] = current_val + curr_id_frag
                         rows_to_drop.append(i)

        if rows_to_drop:
            df = df.drop(rows_to_drop).reset_index(drop=True)

    # 5. Extract Vehicle No from Header Rows
    if "travel_date_time" in df.columns:
        if "vehicle_number" not in df.columns:
            df["vehicle_number"] = None

        current_vehicle = None
        rows_to_drop = []

        for idx, row in df.iterrows():
            val = str(row["travel_date_time"]).strip()
            match = re.search(r'([A-Z]{2}[0-9]{1,2}[A-Z]{0,3}[0-9]{4})', val.replace(" ", ""))
            is_date = re.search(r'\d{2}-\d{2}-\d{4}', val)
            
            if match and not is_date:
                current_vehicle = match.group(1)
                rows_to_drop.append(idx)
            else:
                existing_veh = str(row.get("vehicle_number", "")).strip()
                if current_vehicle and (existing_veh == "" or existing_veh.lower() == "nan"):
                    df.at[idx, "vehicle_number"] = current_vehicle

        if rows_to_drop:
            df = df.drop(rows_to_drop).reset_index(drop=True)

    # 6. Cleaning Helpers
    def _clean_vehicle_no(x):
        return x.replace(" ", "").strip() if isinstance(x, str) else x

    def _clean_datetime(x):
        if not isinstance(x, str): return x
        x = re.sub(r"\s+", " ", x).strip()
        x = re.sub(r"(\d{2})-(\d)\s(\d{3})", r"\1-\2\3", x)
        x = re.sub(r"(\d{2}):(\d)\s(\d):(\d{2})", r"\1:\2\3:\4", x)
        return x

    def _clean_reference_id(x):
        return x.replace(" ", "") if isinstance(x, str) else x

    if "vehicle_number" in df.columns:
        df["vehicle_number"] = df["vehicle_number"].apply(_clean_vehicle_no)
    
    if "travel_date_time" in df.columns:
        df["travel_date_time"] = df["travel_date_time"].apply(_clean_datetime)
    
    if "unique_transaction_id" in df.columns:
        df["unique_transaction_id"] = df["unique_transaction_id"].apply(_clean_reference_id)
        def safe_convert(x):
            try:
                if pd.isna(x): return ""
                x_str = str(x).strip()
                if x_str.replace('.', '', 1).isdigit():
                    return format(float(x_str), ".0f")
                return x_str
            except:
                return str(x)
        df["unique_transaction_id"] = df["unique_transaction_id"].apply(safe_convert)

    if "activity" in df.columns:
        df["activity"] = df["activity"].astype(str).str.strip()
        df = df[~df["activity"].str.lower().isin(["recharge", "", "nan", "none"])]

    # ----------------------------------------------------------------------
    # 🔥 FIX: SMART COLUMN MAPPING (CATCHES PLAZA NAME & ID)
    # ----------------------------------------------------------------------
    final_map = {}
    
    for col in df.columns:
        c = col.lower()
        
        if "vehicle" in c:
            final_map[col] = "Vehicle No"
        elif "date" in c and "time" in c:
            final_map[col] = "Travel Date Time"
        elif "unique" in c or ("transaction" in c and "id" in c):
            final_map[col] = "Unique Transaction ID"
        elif "activity" in c:
            final_map[col] = "Activity"
        elif "debit" in c or "amount" in c:
            final_map[col] = "Tag Dr/Cr"
            
        # --- Plaza Logic ---
        # 1. Plaza ID: Look for "plaza" + "id" OR "lane" + "id"
        elif ("plaza" in c and "id" in c) or ("lane" in c and "id" in c):
             final_map[col] = "Plaza ID"
             
        # 2. Plaza Name: Look for "plaza" (without ID) OR "description" OR "toll"
        elif "plaza" in c or "description" in c or "toll" in c:
             # Ensure we don't accidentally map the ID column if logic overlapped
             if "id" not in c:
                 final_map[col] = "Plaza Name"

    df.rename(columns=final_map, inplace=True)

    return df
# ==========================================
# 4. MAIN FASTAG DATA CLEANER (PDF)
# ==========================================
def process_fastag_data(file_data_list):
    """
    file_data_list: List of tuples -> [(filename, bytes), (filename, bytes)]
    """
    try:
        print(f"🔹 Starting Fastag Processing for {len(file_data_list)} files...")
        
        processed_dfs = []

        for filename, content in file_data_list:
            try:
                fname_lower = filename.lower()
                
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    df_temp = None
                    
                    if "idfc.pdf" in fname_lower:
                        print(f"🔹 File '{filename}' -> Detected IDFC Logic")
                        df_temp = _process_idfc(pdf)
                    
                    elif "icici.pdf" in fname_lower:
                        print(f"🔹 File '{filename}' -> Detected ICICI Logic")
                        df_temp = _process_icici(pdf)
                    
                    else:
                        print(f"⚠️ File '{filename}' -> No Bank Name found. Defaulting to ICICI.")
                        df_temp = _process_icici(pdf)

                    if df_temp is not None and not df_temp.empty:
                        processed_dfs.append(df_temp)

            except Exception as e:
                print(f"⚠️ Error reading file {filename}: {e}")
                continue

        if not processed_dfs:
            print("❌ No valid data extracted.")
            return None, None, None

        # Merge
        final_df = pd.concat(processed_dfs, ignore_index=True)

        # Enforce Columns
        desired_columns = [
            "Vehicle No", "Travel Date Time", "Unique Transaction ID", 
            "Activity", "Plaza Name", "Plaza ID", "Tag Dr/Cr"
        ]

        for col in desired_columns:
            if col not in final_df.columns:
                final_df[col] = ""

        final_df = final_df[desired_columns]

        # General Cleaning (Safety check)
        if "Travel Date Time" in final_df.columns:
            mask = final_df["Travel Date Time"].astype(str).str.lower().str.contains("date|total|page", na=False)
            final_df = final_df[~mask]

        final_df["Vehicle No"] = (
            final_df["Vehicle No"].astype(str)
            .str.replace(" ", "", regex=False)
            .str.replace("-", "", regex=False)
            .str.upper()
            .replace("NAN", "")
            .replace("NONE", "")
        )

        final_df["Tag Dr/Cr"] = pd.to_numeric(
            final_df["Tag Dr/Cr"].astype(str).str.replace(",", ""), errors='coerce'
        ).fillna(0)

        final_df = final_df.fillna("")

        print(f"🔹 Processing complete. Final shape: {final_df.shape}")
        
        from .cleaner import create_styled_excel 
        return create_styled_excel(final_df, "Fastag_Cleaned")

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ Fastag Cleaner Error: {e}")
        return None, None, None