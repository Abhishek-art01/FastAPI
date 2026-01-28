import pandas as pd
import numpy as np
import pdfplumber
import io
import re
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import traceback
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

from .cleaner_helper import (
    get_mandatory_columns, 
    get_xls_style_data, 
    standardize_dataframe, 
    format_excel_sheet,
    clean_columns,
    clean_address
)


# ==========================================
# 1. CLIENT DATA CLEANER
# ==========================================
def process_client_data(file_content):
    try:
        # Columns to Remove based on your request
        DROP_COLS = [
            'Bunit ID', 'Cycle Start', 'Cycle End', 'Billing Period', 'Project', 'Cost Center', 
            'Department', 'Planned Emp Count', 'Travelled Emp Count', 'Billable Emp Count', 
            'No Show', 'Planned Escort', 'Actual Escort', 'Emp Km', 'Trip Cost', 
            'Trip AC Cost', 'Per Emp Cost', 'Escort Cost', 'Penalty', 'Vendor Penalty', 
            'Total Cost', 'Assigned Contract', 'Cab Contract', 'Billing Zone', 
            'Trip Billing Zone', 'Emp Sigin Type', 'Escort ID', 'Toll Cost', 
            'State Tax Cost', 'Parking Or Toll Cost', 'Per Employee Overhead Cost', 
            'Trip Source', 'Extra Kms Based On Billable Employee Count', 'Billing Kms', 
            'Actual Kms At Employee Level', 'Grid Km', 'Employee Adjustment Distance', 
            'Trip Adjustment', 'Total Distance'
        ]

        # Input Mapping
        COL_MAP = {
            "Shift Date": "shift_date", "Trip ID": "trip_id", "Employee ID": "employee_id", 
            "Gender": "gender", "Employee Name": "employee_name", "Shift Time": "shift_time", 
            "Pickup Time": "pickup_time", "Drop Time": "drop_time", "Trip Direction": "trip_direction", 
            "Cab Reg No": "cab_reg_no", "Cab Type": "cab_type", "Vendor": "vendor", 
            "Office": "office", "Airport Name": "airport_name", "Landmark": "landmark", 
            "Address": "address", "Flight Number": "flight_number", "Flight Category": "flight_category", 
            "Flight Route": "flight_route", "Flight Type": "flight_type"
        }
        
        df = pd.read_excel(io.BytesIO(file_content), dtype=str)
        
        # 1. Drop Unwanted
        df = df.drop(columns=[c for c in DROP_COLS if c in df.columns], errors='ignore')

        # 2. Data Cleaning
        if "Cab Reg No" in df.columns:
            df["Cab Reg No"] = df["Cab Reg No"].str.replace("-", "", regex=False).str.upper()

        if "Trip Direction" in df.columns:
            df["Trip Direction"] = df["Trip Direction"].astype(str).str.strip().str.title().replace({
                "Login": "Pickup", "Logout": "Drop"
            })

        # 3. Rename to DB Columns
        df_db = df.rename(columns=COL_MAP).fillna("")
        
        # 4. Standardize (Matches ClientData Model)
        df_db = standardize_dataframe(df_db)
        if df_db is None: return None, None, None

        return create_styled_excel(df_db, "Client_Cleaned")
    except Exception as e:
        print(f"❌ Client Cleaner Error: {e}")
        return None, None, None

# ==========================================
# 2. RAW DATA CLEANER
# ==========================================
def _clean_single_raw_df(df):
    try:

        df = df.replace({np.nan: None, "nan": None})
        # Trip ID Logic
        df["Trip_ID"] = np.where(df.iloc[:, 10].astype(str).str.startswith("T"), df.iloc[:, 10], np.nan)
        df["Trip_ID"] = df["Trip_ID"].ffill()

        # Identify Row Types
        is_header = df.iloc[:, 1].astype(str).str.contains("UNITED FACILITIES", na=False)
        # DEBUG: Relaxed regex slightly to ensure we catch rows even if there are formatting oddities
        is_passenger = df.iloc[:, 0].astype(str).str.match(r"^[0-9]+$") 

        # Extraction Maps
        h_map = {0: 'TRIP_DATE', 1: 'AGENCY_NAME', 2: 'D_LOGIN', 3: 'VEHICLE_NO', 4: 'DRIVER_NAME', 6: 'DRIVER_MOBILE', 7: 'MARSHALL', 8: 'DISTANCE', 9: 'EMP_COUNT', 10: 'TRIP_COUNT'}
        p_map = {0: 'PAX_NO', 1: 'REPORTING_TIME', 2: 'EMPLOYEE_ID', 3: 'EMPLOYEE_NAME', 4: 'GENDER', 5: 'EMP_CATEGORY', 6: 'FLIGHT_NO.', 7: 'ADDRESS', 8: 'REPORTING_LOCATION', 9: 'LANDMARK', 10: 'PASSENGER_MOBILE'}

        df_h = df[is_header].rename(columns=h_map)
        df_p = df[is_passenger].rename(columns=p_map)
        
        # DEBUG CHECK
        if df_h.empty or df_p.empty:
            print("DEBUG: Header or Passenger dataframe is empty. Check regex or file format.")

        cols_h = [c for c in h_map.values() if c in df_h.columns] + ['Trip_ID']
        cols_p = [c for c in p_map.values() if c in df_p.columns] + ['Trip_ID']
        
        merged = pd.merge(df_p[cols_p], df_h[cols_h], on='Trip_ID', how='left')

        # Cleaning Logic
        for col in ['TRIP_ID', 'Trip_ID']:
            if col in merged.columns:
                merged[col] = merged[col].astype(str).str.replace('T', '', regex=False)
                merged.rename(columns={col: 'TRIP_ID'}, inplace=True)

        if 'AGENCY_NAME' in merged.columns:
            merged['AGENCY_NAME'] = merged['AGENCY_NAME'].apply(lambda x: "UNITED FACILITIES" if "UNITED FACILITIES" in str(x).upper() else x)

        if 'VEHICLE_NO' in merged.columns:
            merged['VEHICLE_NO'] = merged['VEHICLE_NO'].astype(str).str.replace('-', '', regex=False)

        if 'D_LOGIN' in merged.columns:
            login = merged['D_LOGIN'].astype(str).str.strip().str.split(' ', n=1, expand=True)
            if len(login.columns) > 0: merged['DIRECTION'] = login[0].str.upper().replace({'LOGIN': 'PICKUP', 'LOGOUT': 'DROP'})
            if len(login.columns) > 1: merged['SHIFT_TIME'] = login[1]

        # Explicit Removal
        cols_to_remove = ['PAX_NO', 'D_LOGIN', 'MARSHALL', 'DISTANCE', 'EMP_COUNT', 'TRIP_COUNT']
        merged = merged.drop(columns=cols_to_remove, errors='ignore')

        for col in merged.select_dtypes(include=['object']):
            merged[col] = merged[col].astype(str).str.upper().str.strip()

        return merged
    except Exception as e:
        # DEBUG: Print actual error
        print(f"Error in _clean_single_raw_df: {e}")
        traceback.print_exc()
        return pd.DataFrame()

def process_raw_data(file_list_bytes):
    all_dfs = []
    for filename, content in file_list_bytes: # Added filename to loop for better debug
        try:
            print(f"Processing file: {filename}") # DEBUG
            df_raw = pd.read_excel(io.BytesIO(content), header=None,dtype=str).dropna(how="all").reset_index(drop=True)
            cleaned = _clean_single_raw_df(df_raw)
            if not cleaned.empty: 
                all_dfs.append(cleaned)
            else:
                print(f"Warning: File {filename} resulted in empty data.")
        except Exception as e:
            # DEBUG: Print actual error
            print(f"FAILED processing file {filename}: {e}")
            traceback.print_exc()
            pass

    if not all_dfs: 
        print("No valid dataframes found in any files.")
        return None, None, None
        
    final_df = pd.concat(all_dfs, ignore_index=True)

    # Input Map -> DB Columns
    DB_MAP = {
        'TRIP_DATE': 'shift_date', 'TRIP_ID': 'trip_id', 'AGENCY_NAME': 'vendor', 
        'FLIGHT_NO.': 'flight_number', 'EMPLOYEE_ID': 'employee_id', 'EMPLOYEE_NAME': 'employee_name', 
        'GENDER': 'gender', 'EMP_CATEGORY': 'emp_category', 'ADDRESS': 'address', 
        'PASSENGER_MOBILE': 'passenger_mobile', 'LANDMARK': 'landmark', 'VEHICLE_NO': 'cab_reg_no',
        'DRIVER_NAME': 'driver_name', 'DRIVER_MOBILE': 'driver_mobile', 'DIRECTION': 'trip_direction',
        'SHIFT_TIME': 'shift_time', 'REPORTING_TIME': 'pickup_time', 'REPORTING_LOCATION': 'office'    
    }

    final_db = final_df.rename(columns=DB_MAP).fillna("")
    
    # 2. Date/Time Formatting
    if 'shift_date' in final_db.columns:
        final_db['shift_date'] = pd.to_datetime(final_db['shift_date'], errors='coerce').dt.strftime('%d-%m-%Y')
        final_db['trip_date'] = final_db['shift_date'] # Duplicate to trip_date
    
    if 'shift_time' in final_db.columns:
        final_db['shift_time'] = pd.to_datetime(final_db['shift_time'], errors='coerce', format='mixed').dt.strftime('%H:%M')

    # 3. Standardize (Matches RawTripData Model)
    # DEBUG: Wrapped in try/except in case this function is missing or failing
    try:
        final_db = standardize_dataframe(final_db)
    except NameError:
        print("CRITICAL: 'standardize_dataframe' function is not defined.")
    except Exception as e:
        print(f"Error in standardize_dataframe: {e}")
        traceback.print_exc()
        return None, None, None

    if final_db is None: return None, None, None

    # DEBUG: Wrapped in try/except
    try:
        return create_styled_excel(final_db, "Raw_Cleaned")
    except NameError:
        print("CRITICAL: 'create_styled_excel' function is not defined.")
        return final_db, None, None
# ==========================================
# 3. OPERATION DATA CLEANER
# ==========================================


def process_operation_data(file_list_bytes):
    # 1. Configuration
    COLUMN_TO_RENAME = {
        'DATE': 'shift_date', 'TRIP ID': 'trip_id', 'FLT NO.': 'flight_number', 
        'SAP ID': 'employee_id', 'EMP NAME': 'employee_name', 'EMPLOYEE ADDRESS': 'employee_address', 
        'PICKUP LOCATION': 'landmark', 'DROP LOCATION': 'office', 'CAB NO': 'cab_last_digit',
        'PICKUP TIME': 'pickup_time', 'REMARKS': 'mis_remark'
    }
    SKIP_HEADERS = ['CONTACT NO', 'GUARD ROUTE', 'AIRPORT DROP TIME']

    wb = Workbook()
    ws = wb.active
    ws.title = "Operation_Data"
    
    # ... (Styles setup same as before) ...
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    try:
        MANDATORY_HEADERS = get_mandatory_columns()
    except:
        MANDATORY_HEADERS = list(COLUMN_TO_RENAME.values())

    for col_idx, header in enumerate(MANDATORY_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    target_row = 2
    data_rows = [] 
    extra_headers_map = {} 
    next_extra_col_idx = len(MANDATORY_HEADERS) + 1

    # 3. Processing Loop
    for filename, content in file_list_bytes:
        if not filename.lower().endswith('.xls'): continue
        print(f"\n--- Processing File: {filename} ---")

        try:
            rb = xlrd.open_workbook(file_contents=content, formatting_info=True)
            rs = rb.sheet_by_index(0)
            source_headers = [str(rs.cell_value(0, c)).strip().upper() for c in range(rs.ncols)]
            
            # Map columns
            # --- START ADDED LOGIC: IDENTIFY SPECIFIC COLUMN INDICES ---
            idx_trip = next((i for i, h in enumerate(source_headers) if 'TRIP ID' in h), None)
            idx_sap = next((i for i, h in enumerate(source_headers) if 'SAP ID' in h), None)
            idx_addr = next((i for i, h in enumerate(source_headers) if 'EMPLOYEE ADDRESS' in h), None)

            col_to_target_map = {}
            # --- END ADDED LOGIC ---

            for idx, raw_header in enumerate(source_headers):
                if any(skip in raw_header for skip in SKIP_HEADERS): continue
                match = next((val for key, val in COLUMN_TO_RENAME.items() if key in raw_header), None)
                if match:
                    col_to_target_map[idx] = {'type': 'mandatory', 'name': match}
                else:
                    if raw_header not in extra_headers_map:
                        extra_headers_map[raw_header] = next_extra_col_idx
                        ws.cell(row=1, column=next_extra_col_idx, value=raw_header)
                        next_extra_col_idx += 1
                    col_to_target_map[idx] = {'type': 'extra', 'name': raw_header}

            # 4. Process Data Rows
            for r_idx in range(1, rs.nrows):
                row_vals = [str(rs.cell_value(r_idx, c)).strip() for c in range(rs.ncols)]
                if sum(1 for v in row_vals if v != "") <= 3: 
                    continue

                row_data_map = {} 
                db_row_dict = {}
                
                # Logic Flags for Style Detection across the entire row
                row_has_yellow_bg = False
                row_has_red_font = False

                # --- START ADDED LOGIC: 3-COLUMN COUNTERS ---
                red_count = 0
                yellow_count = 0
                check_indices = [idx for idx in [idx_trip, idx_sap, idx_addr] if idx is not None]
                # --- END ADDED LOGIC ---

                # Pass 1: Extract data and scan row for color indicators
                for c_idx in range(rs.ncols):
                    bg, fg, is_bold = get_xls_style_data(rb, rs.cell_xf_index(r_idx, c_idx), r_idx, c_idx)
                    
                    # --- START ADDED LOGIC: UPDATE COUNTERS BASED ON 3 SPECIFIC COLUMNS ---
                    if c_idx in check_indices:
                        if fg == "FF0000": red_count += 1
                        if bg == "FFFF00": yellow_count += 1
                    # --- END ADDED LOGIC ---
                    
                    if fg == "FF0000": row_has_red_font = True
                    if bg == "FFFF00": row_has_yellow_bg = True
                    
                    if c_idx in col_to_target_map:
                        target_header = col_to_target_map[c_idx]['name']
                        val = rs.cell_value(r_idx, c_idx)
                        row_data_map[target_header] = {'val': val, 'bg': bg, 'fg': fg, 'bold': is_bold}
                        db_row_dict[target_header] = val

                # --- START ADDED LOGIC: RE-EVALUATE FLAGS BASED ON 3-COLUMN RULE ---
                row_has_red_font = (red_count == 3)
                row_has_yellow_bg = (yellow_count == 3)
                # --- END ADDED LOGIC ---

                # Pass 2: Apply Business Logic Overrides (Priority: Red > Yellow)
                if row_has_red_font:
                    db_row_dict['mis_remark'] = "Cancel"
                    print(f"[LOGIC] Row {r_idx}: Red found -> Marked Cancel")
                elif row_has_yellow_bg:
                    db_row_dict['mis_remark'] = "Alt Veh"
                    print(f"[LOGIC] Row {r_idx}: Yellow found -> Marked Alt Veh")

                # Pass 3: Write to Excel Output
                ws.row_dimensions[target_row].height = 25
                for c_out, m_header in enumerate(MANDATORY_HEADERS, 1):
                    cell = ws.cell(row=target_row, column=c_out)
                    
                    # Get value from dict (contains overrides like 'Cancel')
                    cell.value = db_row_dict.get(m_header, "")
                    
                    # Styling for the mis_remark column based on triggers
                    if m_header == 'mis_remark':
                        if row_has_red_font:
                            cell.font = Font(color="FF0000", bold=True)
                        elif row_has_yellow_bg:
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
                            cell.font = Font(bold=True)
                    elif m_header in row_data_map:
                        # Carry over original formatting for other columns
                        d = row_data_map[m_header]
                        if d['bg']:
                            cell.fill = PatternFill(start_color=d['bg'], end_color=d['bg'], fill_type='solid')
                        cell.font = Font(bold=d['bold'], color=d['fg'] if d['fg'] else None)
                    
                    cell.alignment = align_center
                    cell.border = border

                # Extra check for dynamic extra columns
                for extra_name, extra_col_idx in extra_headers_map.items():
                    cell = ws.cell(row=target_row, column=extra_col_idx)
                    if extra_name in row_data_map:
                        cell.value = row_data_map[extra_name]['val']
                        cell.alignment = align_center
                        cell.border = border

                # Append row to DB list
                if db_row_dict.get('employee_id') or db_row_dict.get('employee_name'):
                    data_rows.append(db_row_dict)
                    target_row += 1
            
            rb.release_resources()
        except Exception as e:
            print(f"[BREAKING ERROR] File {filename}: {e}")
            traceback.print_exc()

    # --- 2. DATAFRAME CLEANING & MATH ---
    df_db = pd.DataFrame(data_rows)
    if not df_db.empty:
        # 1. Helper: Convert Serial Date to DD-MM-YYYY
        def convert_date(d):
            try:
                # Handle Excel Serial (e.g., 46023)
                f_val = float(d)
                # Excel's base date is 1899-12-30
                dt = datetime(1899, 12, 30) + timedelta(days=f_val)
                return dt.strftime('%d-%m-%Y')
            except:
                return str(d)

        # 2. Helper: Convert Serial Time to HH:MM
        def convert_time(t):
            try:
                f_val = float(t) % 1 # MOD 1 logic
                seconds = int(round(f_val * 86400))
                return (datetime.min + timedelta(seconds=seconds)).strftime('%H:%M')
            except:
                return str(t)

        print("[DEBUG] Converting Date to DD-MM-YYYY and calculating Shift Time...")
        df_db['shift_date'] = df_db['shift_date'].apply(convert_date)
        df_db['pickup_time'] = df_db['pickup_time'].apply(convert_time)

        # 3. Logic: SHIFT TIME = PICKUP TIME + 2 HOURS
        # Convert DD-MM-YYYY back to datetime for calculation
        temp_pickup_dt = pd.to_datetime(df_db['shift_date'] + " " + df_db['pickup_time'], dayfirst=True, errors='coerce')
        
        # Add 2 Hours
        temp_shift_dt = temp_pickup_dt + pd.Timedelta(hours=2)
        shift_date_dt = pd.to_datetime(df_db["shift_date"], dayfirst=True).dt.date

        # 4. Populate Final Columns
        df_db['shift_time'] = temp_shift_dt.dt.strftime('%H:%M')
        fixed_drop_dt = temp_shift_dt.where(
            temp_shift_dt.dt.date == shift_date_dt,
            temp_shift_dt - pd.Timedelta(days=1)
        )
        df_db["drop_time"] = fixed_drop_dt.dt.strftime("%d-%m-%Y %H:%M")

        # Keep pickup_time as HH:MM
        pickup_dt = fixed_drop_dt - pd.Timedelta(hours=2)
        df_db["pickup_time"] = pickup_dt.dt.strftime("%d-%m-%Y %H:%M")


        # Step C: FORCE UPPERCASE HEADERS & VALUES
        df_db.columns = df_db.columns.str.strip().str.upper()
        
        # Step D: Numeric Conversion (INT)
        numeric_cols = ["TRIP_ID", "EMPLOYEE_ID", "CAB_LAST_DIGIT"]
        for col in numeric_cols:
            if col in df_db.columns:
                # errors='coerce' turns junk into NaN, then we fill with 0 to allow int conversion
                df_db[col] = pd.to_numeric(df_db[col], errors="coerce").fillna(0).astype(int).astype(str)

        # Step E: Address Cleaning (Regex)
        if "EMPLOYEE_ADDRESS" in df_db.columns:
            df_db["EMPLOYEE_ADDRESS"] = (
                df_db["EMPLOYEE_ADDRESS"]
                .astype(str)
                .str.replace(r"[- , /]", " ", regex=True)
                .str.replace(r"\s+", " ", regex=True)
                .str.strip()
                .str.upper()
            )

        # Step F: Final Global Uppercase for All Text
        text_cols = df_db.select_dtypes(include="object").columns
        df_db[text_cols] = df_db[text_cols].apply(lambda x: x.str.upper())
        df_db = df_db.replace("NAN", "").fillna("")

        # --- 3. WRITE TO EXCEL (FROM CLEANED DATAFRAME) ---
        FINAL_HEADERS = [h.upper() for h in MANDATORY_HEADERS]
        for col_idx, header in enumerate(FINAL_HEADERS, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write Uppercase Headers
        for col_idx, header in enumerate(FINAL_HEADERS, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write Cleaned Values
        for r_idx, row_data in enumerate(df_db.to_dict('records'), 2):
            for c_idx, header in enumerate(MANDATORY_HEADERS, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row_data.get(header, ""))
                cell.alignment = align_center
                cell.border = border

    # Final Styles
    format_excel_sheet(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return df_db, output, "Operation_Cleaned.xlsx"

# ==========================================
# 4. BAROW DATA CLEANER
# ==========================================
def process_ba_row_data(file_content):
    try:
        print("🔹 Starting BA Row Data Processing...")
        
        # 1. READ CSV
        df = pd.read_csv(io.BytesIO(file_content), low_memory=False)
        print(f"🔹 CSV Loaded. Columns: {list(df.columns[:5])}...")

        df.columns = df.columns.str.strip()
        print(f"🔹 CSV Loaded. Found Columns: {list(df.columns)}")

        # 2. FILTER & TRANSFORM
        if "Trip Id" in df.columns:
            df["Trip Id"] = pd.to_numeric(df["Trip Id"], errors='coerce').fillna(0)
            df = df[df["Trip Id"] != 0]

        # Shift Time Logic
        if "Trip Type" in df.columns:
            df["Shift Time"] = df["Trip Type"].astype(str)
            mask_log = df["Shift Time"].str.contains("LOGIN|LOGOUT", case=False, na=False)
            df.loc[mask_log, "Shift Time"] = "00:00"
        else:
            df["Shift Time"] = "00:00"

        # Trip Direction
        if "Direction" in df.columns:
            df["Trip Direction"] = df["Direction"].str.upper().map({
                "LOGIN": "PICKUP", "LOGOUT": "DROP"
            })
        else:
            df["Trip Direction"] = ""

        # Safe Column Access for Pickup/Drop
        df["Pickup Time"] = df.get("Duty Start", "")
        df["Drop Time"] = df.get("Duty End", "")

        # Registration Cleaning
        if "Registration" in df.columns:
            df["Registration"] = (
                df["Registration"].astype(str)
                .str.replace("-", "", regex=False)
                .str.replace(" ", "", regex=False)
                .str.upper()
            )

        # Location Logic
        is_drop = df["Trip Direction"] == "DROP"
        is_pickup = df["Trip Direction"] == "PICKUP"
        
        start_addr = df.get("Start Location Address", "")
        end_addr = df.get("End Location Address", "")
        start_land = df.get("Start Location Landmark", "")
        end_land = df.get("End Location Landmark", "")

        df["Airport Name"] = np.where(is_drop, start_addr, end_addr)
        df["Address"] = np.where(is_pickup, start_addr, end_addr)
        df["Landmark"] = np.where(is_pickup, start_land, end_land)

        if "Leg Date" in df.columns:
            # Standard case
            df["Trip Date"] = df["Leg Date"].astype(str) + " " + df["Shift Time"].astype(str)
        elif "Date" in df.columns:
            # Fallback to 'Date' column
            df["Trip Date"] = df["Date"].astype(str) + " " + df["Shift Time"].astype(str)
        elif "Pickup Time" in df.columns:
            # Fallback: Extract Date from Duty Start (e.g., '2026-01-01 18:00:00')
            print("⚠️ 'Leg Date' missing. Extracting date from 'Pickup Time'.")
            df["Temp_Date"] = pd.to_datetime(df["Pickup Time"], errors='coerce').dt.strftime('%d-%m-%Y')
            df["Trip Date"] = df["Temp_Date"].astype(str) + " " + df["Shift Time"].astype(str)
            df["Leg Date"] = df["Temp_Date"] # Fill Leg Date so it's not empty in DB
        else:
            # Worst case: No date found
            print("❌ ERROR: Could not find 'Leg Date', 'Date', or 'Pickup Time'. Trip Date will be empty.")
            df["Leg Date"] = ""
            df["Trip Date"] = ""

        if "Trip Date" in df.columns:
            df["Shift Date"] = df["Trip Date"]
        else:
            df["Shift Date"] = ""

        df["In App/ Extra"] = "BA Row Data"
        df["BA REMARK"] = df.get("Trip Status", "")
        df["MiS Remark"] = df.get("Comments", "")

        # 3. PREPARE DATABASE MAPPING (Title Case -> snake_case)
        DB_MAP = {
            "Leg Date": "leg_date",
            "Trip Id": "trip_id",
            "Employee ID": "employee_id",
            "Gender": "gender",
            "EMP_CATEGORY": "emp_category",
            "Employee Name": "employee_name",
            "Shift Time": "shift_time",
            "Pickup Time": "pickup_time",
            "Drop Time": "drop_time",
            "Trip Direction": "trip_direction",
            "Registration": "cab_reg_no",
            "Cab Type": "cab_type",
            "Vendor": "vendor",
            "Office": "office",
            "Airport Name": "airport_name",
            "Landmark": "landmark",
            "Address": "address",
            "Flight Number": "flight_number",
            "Flight Category": "flight_category",
            "Flight Route": "flight_route",
            "Flight Type": "flight_type",
            "Trip Date": "trip_date",
            "MiS Remark": "mis_remark",
            "In App/ Extra": "in_app_extra",
            "Traveled Employee Count": "traveled_emp_count",
            "UNA2": "una2",
            "UNA": "una",
            "BA REMARK": "ba_remark",
            "Route Status": "route_status",
            "Clubbing Status": "clubbing_status",
            "GPS TIME": "gps_time",
            "GPS REMARK": "gps_remark",
            "Billing Zone Name": "billing_zone_name",
            "Leg Type": "leg_type",
            "Trip Source": "trip_source",
            "Trip Type": "trip_type",
            "Leg Start": "leg_start",
            "Leg End": "leg_end",
            "Audit Results": "audit_results",
            "Audit Done By": "audit_done_by",
            "Trip Audited": "trip_audited"
        }

        # 4. FIX: ENSURE ALL COLUMNS EXIST
        # This loop prevents the KeyError by creating missing columns
        for col in DB_MAP.keys():
            if col not in df.columns:
                df[col] = ""

        # 5. SELECT AND RENAME
        # Now it is safe to select because we guaranteed they exist
        df_final = df[list(DB_MAP.keys())].copy()
        df_final.rename(columns=DB_MAP, inplace=True)

        print(f"🔹 Data Transformed. Renamed columns to: {list(df_final.columns[:5])}...")
        print("🔹 Calling Standardizer...")

        # 6. STANDARDIZE
        if 'standardize_dataframe' in globals():
            df_final = standardize_dataframe(df_final)
            
            if df_final is None:
                print("❌ Standardization failed. Check if DB model matches these columns.")
                return None, None, None
        else:
            print("⚠️ 'standardize_dataframe' function not found. Skipping.")

        # 7. EXPORT
        print("🔹 Generating Excel...")
        return create_styled_excel(df_final, "BA_Row_Data_Cleaned")

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ BA Row Data Cleaner Error: {e}")
        return None, None, None


