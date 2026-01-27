import pandas as pd
import numpy as np
import pdfplumber
import io
import re
from sqlmodel import Session, select, col
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import traceback
from openpyxl.utils import get_column_letter

#=================================================================
from ..models import OperationData
#=================================================================


#=================================================================

metadata = {'created_at','updated_at','operation_type','processed_by'}

def get_mandatory_columns():
    """Get all column names from OperationData model"""
    columns = []
    
    # Get fields from model
    for field_name, field in OperationData.__fields__.items():
        # Skip metadata columns
        if field_name not in metadata:
            columns.append(field_name)
    
    return columns



def format_excel_headers(ws, start_row=1, start_col=1):
    """
    Format headers in an Excel worksheet.
    
    Args:
        ws: Worksheet object
        start_row: Row number where headers start (default: 1)
        start_col: Column number where headers start (default: 1)
    """
    
    # Define styling
    header_fill = PatternFill(
        start_color="366092",  # Dark blue
        end_color="366092",
        fill_type="solid"
    )
    
    header_font = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="FFFFFF"  # White
    )
    
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True  # Enable text wrapping
    )
    
    # Get max column with data
    max_column = ws.max_column
    
    # Apply formatting to each header cell
    for col in range(start_col, max_column + 1):
        cell = ws.cell(row=start_row, column=col)
        
        # Apply styling if cell has a value
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    # Set row height for header row
    ws.row_dimensions[start_row].height = 30
    
    # Auto-fit column widths
    for col in range(start_col, max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        # Check all cells in the column (including header)
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            
            if cell.value:
                # Calculate length considering line breaks for wrapped text
                if isinstance(cell.value, str):
                    # Find the longest line if text wraps
                    lines = str(cell.value).split('\n')
                    line_length = max(len(line) for line in lines)
                else:
                    line_length = len(str(cell.value))
                
                # Add a little padding
                adjusted_length = line_length + 2
                
                if adjusted_length > max_length:
                    max_length = adjusted_length
        
        # Set column width (minimum 10, maximum 50)
        column_width = min(max(max_length, 10), 50)
        ws.column_dimensions[column_letter].width = column_width

    return ws


def clean_columns(columns):
    cleaned = (
        columns
        .str.replace(r"\n", " ", regex=True)        # remove line breaks
        .str.replace(r"\t", " ", regex=True)        # remove tabs
        .str.replace(r"\s+", " ", regex=True)       # normalize spaces
        .str.strip()                                # trim edges
        .str.lower()                                # lowercase
        .str.replace(r"[^\w\s]", "", regex=True)    # remove special chars
        .str.replace(" ", "_")                      # snake_case
    )
    return cleaned


def clean_address(series: pd.Series) -> pd.Series:
    """
    Cleans address text exactly like:
    =UPPER(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(A2,"-"," "),","," "),"/"," "))
    """
    return (
        series
        .astype(str)
        .str.replace("-", " ", regex=False)
        .str.replace(",", " ", regex=False)
        .str.replace("/", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)  # normalize spaces
        .str.strip()
        .str.upper()
    )



def format_excel_sheet(ws):
    """
    Final Excel formatter:
    - Header style (blue, bold, white)
    - Cambria font for all cells
    - Wrap text ON
    - Row height = 30
    - Auto-fit columns
    - EMPLOYEE ADDRESS width = 80
    - EMPLOYEE NAME width = 30
    """

    # Styles
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(name="Cambria", size=12, bold=True, color="FFFFFF")
    cell_font = Font(name="Cambria")

    align_center_wrap = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # -----------------------------
    # Header formatting
    # -----------------------------
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center_wrap
        cell.border = border

    ws.row_dimensions[1].height = 30

    # -----------------------------
    # Cell formatting + row height
    # -----------------------------
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 30
        for cell in row:
            cell.font = cell_font
            cell.alignment = align_center_wrap
            cell.border = border

    # -----------------------------
    # Auto-fit column width
    # -----------------------------
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # -----------------------------
    # Override specific columns
    # -----------------------------
    for cell in ws[1]:
        if cell.value == "EMPLOYEE ADDRESS":
            ws.column_dimensions[cell.column_letter].width = 80
        elif cell.value == "EMPLOYEE NAME":
            ws.column_dimensions[cell.column_letter].width = 30



def standardize_dataframe(df):
    """
    Standardizes DataFrame to match SQLModel definitions:
    1. Generates 'unique_id' (Composite Key).
    2. Ensures ALL mapped columns exist (fills missing with "").
    3. Returns DataFrame with [Mandatory Columns] + [Any Extra Columns].
    """
    if df is None or df.empty:
        return None

    # 1. Generate Unique ID
    if 'trip_id' in df.columns and 'employee_id' in df.columns:
        df['unique_id'] = df['trip_id'].astype(str).str.strip() + df['employee_id'].astype(str).str.strip()
        # Filter invalid IDs
        df = df[
            df["unique_id"].notna() & 
            (df["unique_id"] != "") & 
            (~df["unique_id"].str.contains("nan|None", case=False))
        ]
    else:
        return None # Critical missing data

    # 2. Sync Columns with SQLModel (Fill Missing)
    target_cols = list(MANDATORY_DB_MAP.values())
    for col in target_cols:
        if col not in df.columns:
            df[col] = ""

    # 3. Populate specific logic columns
    df['una2'] = df['unique_id'] # As requested
    # Optional: populate 'una' if needed, or leave blank

    # 4. Organize Columns: Mandatory -> Extra -> unique_id
    current_cols = df.columns.tolist()
    reserved_cols = set(target_cols) | {'unique_id'}
    extra_cols = [c for c in current_cols if c not in reserved_cols]
    
    final_order = target_cols + extra_cols + ['unique_id']
    return df[final_order]

def get_xls_style_data(book, xf_index):
    """ Helper to extract style (Hex Color, Bold) from xlrd book. """
    try:
        xf = book.xf_list[xf_index]
        font = book.font_list[xf.font_index]
        
        def get_hex(idx):
            if idx is not None and idx < 64:
                rgb = book.colour_map.get(idx)
                if rgb and rgb != (0,0,0) and rgb != (255, 255, 255):
                    return "{:02x}{:02x}{:02x}".format(*rgb).upper()
            return None

        bg_hex = get_hex(xf.background.pattern_colour_index)
        font_hex = get_hex(font.colour_index)
        return bg_hex, font_hex, bool(font.bold)
    except:
        return None, None, False

def create_styled_excel(df, filename_prefix="Cleaned"):
    """ Generates Excel with consistent formatting. """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Invert map to get Friendly Headers
        inv_map = {v: k for k, v in MANDATORY_DB_MAP.items()}
        df_export = df.rename(columns=inv_map)
        
        # Order: Mandatory Headers -> Extra Headers
        export_cols = [h for h in MANDATORY_HEADERS if h in df_export.columns]
        remaining = [c for c in df_export.columns if c not in MANDATORY_HEADERS and c != 'unique_id']
        export_cols += remaining
        
        df_export = df_export[export_cols]
        
        df_export.to_excel(writer, index=False, sheet_name='Raw_Data')
        
        # Apply Styles
        workbook = writer.book
        worksheet = writer.sheets['Raw_Data']
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#0070C0', 'font_color': 'white'})
        
        for i, col in enumerate(df_export.columns):
            worksheet.write(0, i, col, header_fmt)
            worksheet.set_column(i, i, 15)
            
    output.seek(0)
    return df, output, f"{filename_prefix}.xlsx"

# --- HELPER FUNCTIONS ---
def bulk_save_unique(session: Session, model_class, df: pd.DataFrame, unique_col: str = "unique_id") -> int:
    """Helper to insert only new rows into database based on a unique column."""
    if df is None or df.empty or unique_col not in df.columns:
        return 0
    
    incoming_ids = df[unique_col].dropna().unique().tolist()
    if not incoming_ids: return 0

    existing_ids = set(session.exec(select(getattr(model_class, unique_col)).where(col(getattr(model_class, unique_col)).in_(incoming_ids))).all())
    new_rows = df[~df[unique_col].isin(existing_ids)]
    
    if not new_rows.empty:
        records = [model_class(**row.where(pd.notnull(row), None).to_dict()) for _, row in new_rows.iterrows()]
        session.add_all(records)
        session.commit()
        return len(new_rows)
    return 0

def sync_addresses_to_t3(session: Session, df: pd.DataFrame) -> int:
    """
    Extracts unique addresses from the uploaded dataframe and adds NEW ones
    to the t3_address_locality table. Handles duplicates and ID conflicts.
    """
    # 1. Identify Address Column
    address_col = None
    possible_names = ["address", "Address", "employee_address", "Employee Address", "pickup_location", "drop_location"]
    
    for col_name in possible_names:
        if col_name in df.columns:
            address_col = col_name
            break
    
    if not address_col:
        print("⚠️ T3 Sync: No address column found in uploaded file.")
        return 0

    # 2. Extract Unique Addresses from File
    file_addresses = set(
        df[address_col]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
    )
    file_addresses.discard("") # Remove empty strings

    if not file_addresses:
        return 0

    # 3. Find Addresses ALREADY in Database
    # Fetch all existing addresses to compare against
    # Use chunking if you have millions of rows, but for thousands, this is fine.
    existing_db_addresses = set(session.exec(select(T3AddressLocality.address)).all())
    
    # 4. Filter New Addresses
    new_addresses_list = list(file_addresses - existing_db_addresses)

    if not new_addresses_list:
        print("✅ T3 Sync: All addresses already exist.")
        return 0

    print(f"📍 T3 Sync: Found {len(new_addresses_list)} NEW addresses. Inserting...")

    # 5. Insert One-by-One to Isolate Failures (Safer for debugging conflicts)
    # Or Bulk Insert if confident. Let's try a safer Bulk approach.
    
    records = [T3AddressLocality(address=addr, locality=None) for addr in new_addresses_list]
    
    try:
        session.add_all(records)
        session.commit()
        return len(records)
    except Exception as e:
        session.rollback()
        print(f"❌ T3 Sync Error: {e}")
        
        # AUTO-FIX: Attempt to reset the ID sequence if it's a Primary Key error
        if "t3_address_locality_pkey" in str(e) or "UniqueViolation" in str(e):
            print("🔧 Attempting to fix ID sequence...")
            try:
                # This SQL command resets the ID counter to the max ID + 1
                session.exec(text("SELECT setval(pg_get_serial_sequence('t3_address_locality', 'id'), coalesce(max(id),0) + 1, false) FROM t3_address_locality;"))
                session.commit()
                
                # Retry Insert
                print("🔄 Retrying insert after sequence fix...")
                session.add_all(records)
                session.commit()
                return len(records)
            except Exception as retry_e:
                print(f"❌ Retry Failed: {retry_e}")
                return 0
        return 0
